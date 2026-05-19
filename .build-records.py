"""
일자별 record JSON 생성/갱신 도구.

사용법:
  py .build-records.py                 # 5월 전체 상품별 엑셀 일괄 처리
  py .build-records.py 2026-05-09      # 특정 날짜만 처리
  py .build-records.py 2026-05         # 그 달 전체

흐름:
  1. 상품별매출_EXCEL\상품별_매출_*_{N}일.xlsx → posRows로 변환 (메뉴별 net 매출)
  2. 합계 행에서 카드/현금/외상/기타/할인 추출 → totals 필드 (VAT은 10% 가정으로 추정)
  3. 기존 라벨링_정리본\{date}.json 있으면 purchaseRows·cardRows 보존하고 매출만 갱신
  4. salesMemo에 출처·총매출(서비스/할인 전)·할인합계 메모

일일정산내역 JPG가 있으면 cardRows + 정확한 supply/vat는 별도 단계에서 덮어씀.
"""
import os, sys, json, re
from pathlib import Path
from openpyxl import load_workbook

DRIVE_ROOT = Path(os.environ.get("DRIVE_ROOT", r"G:\내 드라이브\아스트라"))
XLSX_DIR = DRIVE_ROOT / "매출" / "상품별매출_EXCEL"
JSON_DIR = DRIVE_ROOT / "라벨링_정리본"
DEFAULT_MONTH = "2026-05"

# 컬럼 (0-based, 빈 col 0 다음)
C_NO, C_CATEGORY, C_PRODCODE, C_PRODNAME = 1, 2, 3, 4
C_QTY, C_SALES, C_CASH, C_CARD, C_POINT, C_CREDIT, C_ETC, C_DISCOUNT = 5, 6, 7, 8, 9, 10, 11, 12

HEADER_ROW = 3
TOTAL_ROW = 4
DATA_START = 5


def find_xlsx_for_day(day):
    """5월 N일 → 매칭되는 xlsx Path 또는 None (여러 개면 가장 최신)"""
    pat = re.compile(rf'.*_{day}일\.xlsx$')
    candidates = [f for f in XLSX_DIR.glob("*.xlsx") if pat.match(f.name)]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _i(v):
    """정수 변환 (실패 시 0)"""
    if v is None: return 0
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (ValueError, TypeError):
        return 0


def parse_xlsx(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # 합계 행
    total_row = next(ws.iter_rows(min_row=TOTAL_ROW, max_row=TOTAL_ROW, values_only=True))
    qty_total = _i(total_row[C_QTY])
    pos_total = _i(total_row[C_SALES])
    cash_total = _i(total_row[C_CASH])
    card_total = _i(total_row[C_CARD])
    point_total = _i(total_row[C_POINT])
    credit_total = _i(total_row[C_CREDIT])
    etc_total = _i(total_row[C_ETC])
    discount_total = _i(total_row[C_DISCOUNT])
    pos_etc = point_total + credit_total + etc_total  # 카드/현금 외 합

    # 상품 행
    pos_rows = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if row[C_NO] is None:
            continue
        category = row[C_CATEGORY] or ""
        product = row[C_PRODNAME] or ""
        qty = _i(row[C_QTY])
        amount = _i(row[C_SALES])
        # 결제수단: 행마다 cash/card 중 큰 쪽으로 (보통 한쪽이 dominant)
        rc = _i(row[C_CASH])
        rd = _i(row[C_CARD])
        method = "현금" if rc > rd else "카드"
        pos_rows.append({
            "time": "",
            "menu": f"[{category}] {product}",
            "qty": qty,
            "amount": amount,
            "method": method
        })

    # VAT 10% 가정 추정 (일일정산내역 JPG로 나중에 정확값 덮어쓰기 가능)
    supply = round(pos_total / 1.1)
    vat = pos_total - supply

    return {
        "posRows": pos_rows,
        "totals": {
            "posTotal": pos_total,
            "posCard": card_total,
            "posCash": cash_total,
            "posEtc": pos_etc,
            "cardTotal": card_total,
            "supply": supply,
            "vat": vat
        },
        "discount_total": discount_total,
        "qty_total": qty_total,
        "source_file": path.name
    }


def merge_into_json(date_str, parsed):
    """라벨링_정리본\{date}.json 읽고 매출 갱신 후 저장"""
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    json_path = JSON_DIR / f"{date_str}.json"
    if json_path.exists():
        data = json.loads(json_path.read_text(encoding="utf-8"))
    else:
        data = {
            "date": date_str,
            "totals": {},
            "posRows": [],
            "cardRows": [],
            "purchaseRows": [],
            "bankRows": [],
            "extras": {"vatRate": 10, "expenses": 0},
            "issues": [],
            "status": "draft"
        }
    data["date"] = date_str
    data["totals"] = parsed["totals"]
    data["posRows"] = parsed["posRows"]
    data.setdefault("extras", {})["vatRate"] = 10
    # salesMemo: 매출 출처 정보 (cardRows·정확 VAT 없을 수 있음)
    pt = parsed["totals"]["posTotal"]
    disc = parsed["discount_total"]
    data["salesMemo"] = (
        f"POS 상품별 매출 자동 변환 ({parsed['source_file']}, {parsed['qty_total']}개 메뉴 / {len(parsed['posRows'])}품목). "
        f"순매출 {pt:,}원 = 카드 {parsed['totals']['posCard']:,} + 현금 {parsed['totals']['posCash']:,} + 기타 {parsed['totals']['posEtc']:,}. "
        f"할인합계 {disc:,}원. VAT는 10% 가정 추정 (정확값은 일일정산내역 JPG로 덮어쓰기)."
    )
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


def process_day(date_str):
    """date_str: 'YYYY-MM-DD'"""
    day = int(date_str.split("-")[2])
    xlsx = find_xlsx_for_day(day)
    if xlsx is None:
        print(f"  {date_str}: 상품별 엑셀 없음")
        return False
    parsed = parse_xlsx(xlsx)
    json_path = merge_into_json(date_str, parsed)
    pt = parsed["totals"]["posTotal"]
    print(f"  {date_str}: {len(parsed['posRows'])}품목, 순매출 {pt:,}원 → {json_path.name}")
    return True


def process_month(month):
    """month: 'YYYY-MM'. 폴더의 모든 N일 파일 처리."""
    pat = re.compile(r'.*_(\d+)일\.xlsx$')
    days = set()
    for f in XLSX_DIR.glob("*.xlsx"):
        m = pat.match(f.name)
        if m:
            days.add(int(m.group(1)))
    if not days:
        print(f"  {XLSX_DIR}에 상품별 엑셀 없음")
        return
    print(f"  {month} {len(days)}일치 발견: {sorted(days)}")
    for day in sorted(days):
        date_str = f"{month}-{day:02d}"
        process_day(date_str)


def main():
    args = sys.argv[1:]
    if not args:
        target = DEFAULT_MONTH
    else:
        target = args[0]

    if re.match(r'^\d{4}-\d{2}-\d{2}$', target):
        process_day(target)
    elif re.match(r'^\d{4}-\d{2}$', target):
        process_month(target)
    else:
        print(f"Usage: py .build-records.py [YYYY-MM-DD | YYYY-MM]")
        sys.exit(1)


if __name__ == "__main__":
    main()
